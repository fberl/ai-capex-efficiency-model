"""AI_Capex_Efficiency — $ value of cutting AI memory 100x and compute x9.24 (TODAY) / x22.1 (CEILING).

Layout:
  - Totals      : front page, all-company roll-up (live) + GLOBAL estimate row
  - one tab per company (one per entry in ai_capex_model.COMPANIES): full
    bottom-up build  total capex (DISCLOSED) -> infra share -> server bucket ->
    accelerator capex -> fleet/opex -> efficient version -> value, FY25 + FY26
  - Inputs      : global assumptions + the reduction engine (Amdahl cost-weighting)
  - Sensitivity : SpaceX value across reduction tiers
  - CostLadder  : own-silicon vs buy-NVIDIA vs rent $/GPU-hr
  - ServingTraining : the 2026-08-07/08 multi-GPU receipts priced on the ladder
    ($/1M generated tokens vs context; training GPU-hour ratio; assumptions table)
  - Evidence    : BOM split, accelerator-share data, own-silicon TCO, filing top-lines
  - Methodology : steps, caveats, sources

Engine: a GPU is ~60% memory / ~40% compute by cost. TODAY (default): memory x100
+ FLOPs x9.24 (measured blend x2.19 x fit-derived equal-quality parameter ratio
x4.22 at 1T) -> ~20x cost-weighted. CEILING: FLOPs x22.1 (measured prefill x4.02
x 5.5 kernel-campaign speedup) -> ~41x. Floored by compute. NOT 100x.

Every output is a LIVE FORMULA. Colors: yellow = assumption (a lever; each maps to a
slider in the Streamlit app), green = disclosed filing/market data, blue = derived.
Run:  uv run --with openpyxl python ai_capex_efficiency.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ai_capex_model import GLOBALS, COMPANIES  # single source of truth for defaults

INPUT_FILL = PatternFill(
    "solid", fgColor="FFF2CC"
)  # yellow = ASSUMPTION (a lever / app slider)
DATA_FILL = PatternFill(
    "solid", fgColor="E2EFDA"
)  # green  = disclosed filing / market data
CALC_FILL = PatternFill("solid", fgColor="DDEBF7")  # blue   = derived formula
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="BDD7EE")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def safe_text(v):
    """Notes that DESCRIBE a formula start with '= ' (equals+space); strip the
    prefix so Excel stores them as text. Real formulas start with '='+non-space
    (e.g. '=B2*B3') and pass through untouched."""
    if isinstance(v, str) and v.startswith("= "):
        return v[2:]
    return v


def header(ws, row, text, span=6):
    ws.cell(row=row, column=1, value=safe_text(text)).font = WHITE_BOLD
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = HEAD_FILL


def put(ws, r, c, val, fmt=None, fill=None, border=False, bold=False, wrap=False):
    cell = ws.cell(
        row=r, column=c, value=safe_text(val) if isinstance(val, str) else val
    )
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if bold:
        cell.font = BOLD
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# global Inputs addresses (fixed by the input order below)
MEMFAC, FLOPFAC, MEMSHARE = "Inputs!$B$2", "Inputs!$B$3", "Inputs!$B$4"
OPXRED, DISC, GPUCOST = "Inputs!$B$5", "Inputs!$B$6", "Inputs!$B$7"
PWR, ELEC, OH, LIFE = "Inputs!$B$8", "Inputs!$B$9", "Inputs!$B$10", "Inputs!$B$11"
MCAP, DCSCALE, RED = "Inputs!$B$12", "Inputs!$B$13", "Inputs!$B$20"

# uniform key cells on every company tab (col B = FY2025, col C = FY2026)
K_TOTAL25, K_TOTAL26 = "$B$3", "$C$3"
K_ACCEL25, K_ACCEL26 = "$B$12", "$C$12"
K_PCT25 = "$B$13"
K_AVOID25, K_AVOID26 = "$B$25", "$C$25"
K_CAP25, K_CAP26 = "$B$28", "$C$28"
K_AIREV, K_AICAPEX, K_AIOPEX = "$B$32", "$B$33", "$B$34"
K_AINOW, K_AICUT, K_AIARCH, K_AIPCT = "$B$35", "$B$36", "$B$37", "$B$38"


def build_inputs(inp):
    widths(inp, {"A": 40, "B": 13, "C": 8, "D": 76})
    header(
        inp,
        1,
        "GLOBAL INPUTS  —  edit yellow (assumption) cells; green = disclosed/market data",
        span=4,
    )
    G = GLOBALS  # values single-sourced from ai_capex_model; order fixes cell addresses B2..B14 (named-share = B14, used by the Totals global row)
    inputs = [
        (
            "Memory reduction factor",
            G["mem_factor"],
            "x",
            "1e-2 memory = 100x less (RNN O(1) state vs transformer O(T) KV cache).",
            INPUT_FILL,
        ),
        (
            "FLOPs reduction factor",
            # LIVE: TODAY scenario = the ServingTraining workload blend x the
            # equal-quality parameter ratio (Inputs rows 23-25). Overtype with
            # =B25 for CEILING, or any number to pin the lever.
            "=$B$24",
            "x",
            "TODAY scenario (default) = x9.24: the measured workload blend x2.19 (10:1 input:output, "
            "64k context; prefill x1.17, decode x2.20 on the memory ceiling; training excluded) x the "
            "fit-derived equal-quality parameter ratio x4.22 at 1T (rows 23-25 below). Overtype with "
            "=B25 for the CEILING scenario (x22.1 = measured prefill x4.02 x 5.5 kernel speedup), or "
            "any number. See ServingTraining for the blend.",
            INPUT_FILL,
        ),
        (
            "Memory share of GPU cost (BOM)",
            G["mem_share"],
            "frac",
            "HBM ~45% of B200 COGS + most CoWoS packaging -> ~60% memory / ~40% compute. See Evidence.",
            INPUT_FILL,
        ),
        (
            "Opex / energy reduction factor",
            "=B20",
            "x",
            "DERIVED default = cost-weighted reduction (B20): energy splits memory/compute like cost. Overtype this cell to override.",
            INPUT_FILL,
        ),
        (
            "Discount rate",
            G["discount_rate"],
            "frac",
            "Perpetuity capitalization: value = annual benefit / rate (=16.7x at 6%).",
            INPUT_FILL,
        ),
        (
            "Fully-loaded cost per GPU",
            G["gpu_cost"],
            "$",
            "B200-class GPU + share of server, NVLink, networking.",
            INPUT_FILL,
        ),
        (
            "Wall power per GPU",
            G["wall_power_kw"],
            "kW",
            "~1 kW TDP + node overhead x PUE 1.3.",
            INPUT_FILL,
        ),
        (
            "Electricity rate",
            G["elec_rate"],
            "$/kWh",
            "Datacenter wholesale; raise to 0.10-0.12 for grid colo.",
            INPUT_FILL,
        ),
        (
            "Cooling/ops overhead on energy",
            G["cooling_overhead"],
            "frac",
            "Non-power running cost as fraction of electricity.",
            INPUT_FILL,
        ),
        (
            "Fleet useful life",
            G["fleet_life_yr"],
            "yr",
            "AI-GPU depreciation life.",
            INPUT_FILL,
        ),
        (
            "SpaceX market cap",
            G["spacex_mktcap"],
            "$B",
            "~$1.84T Aug 2026; IPO 2026-06-12 at ~$1.77T (market data; used by Sensitivity).",
            DATA_FILL,
        ),
        (
            "Datacenter scaling factor",
            G["dc_scale"],
            "frac",
            "TOGGLE 0-1. 0 = conservative (only accelerator silicon shrinks). 1 = whole AI datacenter (building/power/cooling/net) scales with the smaller fleet. ~0.7 ~= breakeven on net AI.",
            INPUT_FILL,
        ),
        (
            "Named share of global AI capex",
            G["named_share_of_global"],
            "frac",
            "GLOBAL estimate (Totals): named firms' share of worldwide AI capex; the rest (other clouds, China, neoclouds, xAI, sovereign) is grossed up pro-rata. ESTIMATE.",
            INPUT_FILL,
        ),
    ]
    r = 2
    for label, val, unit, note, fill in inputs:
        put(inp, r, 1, label)
        put(
            inp,
            r,
            2,
            val,
            fmt=("0.0%" if unit == "frac" else "#,##0" if unit == "$" else "0.0"),
            fill=fill,
            border=True,
        )
        put(inp, r, 3, unit)
        put(inp, r, 4, note, wrap=True)
        r += 1
    header(inp, 15, "REDUCTION ENGINE (Amdahl cost-weighting) — derived", span=4)
    put(inp, 16, 1, "Compute share of GPU cost")
    put(inp, 16, 2, "=1-B4", fmt="0.0%", fill=CALC_FILL, border=True)
    put(inp, 16, 4, "1 - memory share", wrap=True)
    put(inp, 17, 1, "Memory cost fraction after reduction")
    put(inp, 17, 2, "=B4/B2", fmt="0.00%", fill=CALC_FILL, border=True)
    put(inp, 17, 4, "memory share / memory reduction", wrap=True)
    put(inp, 18, 1, "Compute cost fraction after reduction")
    put(inp, 18, 2, "=B16/B3", fmt="0.00%", fill=CALC_FILL, border=True)
    put(inp, 18, 4, "compute share / FLOPs reduction", wrap=True)
    put(inp, 19, 1, "Residual cost fraction")
    put(inp, 19, 2, "=B17+B18", fmt="0.0%", fill=CALC_FILL, border=True)
    put(inp, 19, 4, "Amdahl: floored by the least-reduced component", wrap=True)
    put(inp, 20, 1, "COST-WEIGHTED reduction factor", bold=True)
    put(inp, 20, 2, "=1/B19", fmt="0.0", fill=CALC_FILL, border=True, bold=True)
    put(inp, 20, 3, "x")
    put(
        inp,
        20,
        4,
        "1 / residual. The realistic $ reduction used everywhere.",
        wrap=True,
    )
    header(inp, 22, "SCENARIOS (2026-08-17 ruling) — TODAY (default) vs CEILING", span=4)
    put(inp, 23, 1, "Equal-quality parameter ratio at 1T")
    put(inp, 23, 2, 4.2185, fmt="0.00", fill=INPUT_FILL, border=True)
    put(inp, 23, 3, "x")
    put(inp, 23, 4,
        "FIT-DERIVED (sealed 2026-08-14 refit): bAttention matches the transformer fit's quality on "
        "23.7% of the params at 1T -> x4.22 compute per token at equal quality. 84.2% at 1B, 55.2% at "
        "10B, 36.2% at 100B. A projection of the two fits, not a measurement.", wrap=True)
    put(inp, 24, 1, "TODAY compute lever = blend x ratio", bold=True)
    put(inp, 24, 2, "=ServingTraining!$B$28*$B$23", fmt="0.00", fill=CALC_FILL, border=True, bold=True)
    put(inp, 24, 3, "x")
    put(inp, 24, 4, "B3 (the FLOPs lever) points here by default.", wrap=True)
    put(inp, 25, 1, "CEILING compute lever")
    put(inp, 25, 2, 22.09, fmt="0.00", fill=INPUT_FILL, border=True)
    put(inp, 25, 3, "x")
    put(inp, 25, 4,
        "Measured prefill x4.02 (bf16 parameter-matched, d512/1M, 2026-07-21) x 5.5 kernel-campaign "
        "speedup. Overtype B3 with =B25 to run the workbook at the ceiling (~41x cost-weighted).",
        wrap=True)


def build_company(ws, name, c25, c26, mcap, ai_rev, basis, sources=()):
    """c25/c26 = (total_capex, infra_share, server_share, accel_share) for FY25/FY26.
    ai_rev = (revenue_FY25, revenue_FY26). sources = [(label, url), ...]."""
    widths(ws, {"A": 34, "B": 12, "C": 12, "D": 54})
    header(
        ws,
        1,
        f"{name} — accelerator capex & value (FY2025 actual + FY2026 estimate)",
        span=4,
    )
    put(ws, 2, 1, "Metric", bold=True)
    put(ws, 2, 2, "FY2025", bold=True)
    put(ws, 2, 3, "FY2026", bold=True)
    put(ws, 2, 4, "basis / source", bold=True)
    t25, i25, s25, a25 = c25
    t26, i26, s26, a26 = c26
    # FY2025 total capex is DISCLOSED (green); FY2026 is an estimate; shares are assumptions (yellow)
    inrows = [
        (3, "Total capex ($B)", t25, t26, "#,##0.0", basis, DATA_FILL, INPUT_FILL),
        (
            4,
            "Infra / data-center share",
            i25,
            i26,
            "0%",
            "strips non-AI (Amazon = AWS only)",
            INPUT_FILL,
            INPUT_FILL,
        ),
        (
            5,
            "Server / short-lived share",
            s25,
            s26,
            "0%",
            "CFO-disclosed (MSFT rose 50%->67%)",
            INPUT_FILL,
            INPUT_FILL,
        ),
        (
            6,
            "Accelerator share within servers",
            a25,
            a26,
            "0%",
            "from BOM teardowns (~67-80%)",
            INPUT_FILL,
            INPUT_FILL,
        ),
    ]
    for r, lab, v25, v26, fmt, note, f25, f26 in inrows:
        put(ws, r, 1, lab)
        put(ws, r, 2, v25, fmt=fmt, fill=f25, border=True)
        put(ws, r, 3, v26, fmt=fmt, fill=f26, border=True)
        put(ws, r, 4, note, wrap=True)
    put(ws, 7, 1, "Market cap ($B, approx)")
    put(ws, 7, 2, mcap, fmt="#,##0", fill=DATA_FILL, border=True)
    put(ws, 7, 4, "approximate June 2026 market data -- edit", wrap=True)

    header(ws, 9, "DERIVATION", span=4)
    der = [
        (
            10,
            "AI-infra capex ($B)",
            "=B3*B4",
            "=C3*C4",
            "#,##0.0",
            "= total x infra share",
            False,
        ),
        (
            11,
            "Server bucket ($B)",
            "=B10*B5",
            "=C10*C5",
            "#,##0.0",
            "= infra x server share",
            False,
        ),
        (
            12,
            "ACCELERATOR capex ($B)",
            "=B11*B6",
            "=C11*C6",
            "#,##0.0",
            "= server x accel share",
            True,
        ),
        (
            13,
            "Accel % of total capex",
            "=B12/B3",
            "=C12/C3",
            "0%",
            "varies across companies",
            False,
        ),
    ]
    for r, lab, bf, cf, fmt, note, bold in der:
        put(ws, r, 1, lab, bold=bold)
        put(ws, r, 2, bf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 3, cf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 4, note, wrap=True)

    header(ws, 15, "FLEET & OPERATING COST (from accelerator capex)", span=4)
    fl = [
        (
            16,
            "Fleet size (GPU-equiv)",
            f"=B12*1000000000/{GPUCOST}",
            f"=C12*1000000000/{GPUCOST}",
            "#,##0",
            "= accel capex / $ per GPU",
        ),
        (
            17,
            "Total wall power (MW)",
            f"=B16*{PWR}/1000",
            f"=C16*{PWR}/1000",
            "#,##0",
            "= GPUs x kW",
        ),
        (18, "Daily energy (MWh)", "=B17*24", "=C17*24", "#,##0", ""),
        (
            19,
            "Daily all-in opex ($/day)",
            f"=B18*1000*{ELEC}*(1+{OH})",
            f"=C18*1000*{ELEC}*(1+{OH})",
            "#,##0",
            "= MWh x rate x (1+overhead)",
        ),
        (20, "Annual opex ($M)", "=B19*365/1000000", "=C19*365/1000000", "#,##0.0", ""),
        (
            21,
            "Lifetime opex ($B)",
            f"=B20*{LIFE}/1000",
            f"=C20*{LIFE}/1000",
            "0.00",
            "",
        ),
    ]
    for r, lab, bf, cf, fmt, note in fl:
        put(ws, r, 1, lab)
        put(ws, r, 2, bf, fmt=fmt, fill=CALC_FILL, border=True)
        put(ws, r, 3, cf, fmt=fmt, fill=CALC_FILL, border=True)
        put(ws, r, 4, note, wrap=True)

    header(ws, 23, "EFFICIENT VERSION & VALUE", span=4)
    va = [
        (
            24,
            "Efficient AI capex ($B)",
            "=B10-B25",
            "=C10-C25",
            "#,##0.0",
            "= AI capex - avoided",
            False,
        ),
        (
            25,
            "Capex avoided/yr ($B)",
            f"=(B12+(B10-B12)*{DCSCALE})*(1-1/{RED})",
            f"=(C12+(C10-C12)*{DCSCALE})*(1-1/{RED})",
            "#,##0.0",
            "accel + (datacenter x dc-scale), all x (1-1/reduction)",
            False,
        ),
        (
            26,
            "Annual opex savings ($M)",
            f"=B20*(1-1/{OPXRED})",
            f"=C20*(1-1/{OPXRED})",
            "#,##0.0",
            "",
            False,
        ),
        (
            27,
            "Sustained annual benefit ($B/yr)",
            "=B25+B26/1000",
            "=C25+C26/1000",
            "#,##0.00",
            "= avoided + opex savings",
            False,
        ),
        (
            28,
            "Capitalized value ($B)",
            f"=B27/{DISC}",
            f"=C27/{DISC}",
            "#,##0",
            "= benefit / discount rate",
            True,
        ),
        (29, "% of market cap", "=B28/$B$7", "=C28/$B$7", "0.0%", "", False),
    ]
    for r, lab, bf, cf, fmt, note, bold in va:
        put(ws, r, 1, lab, bold=bold)
        put(ws, r, 2, bf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 3, cf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 4, note, wrap=True)

    header(ws, 31, "AI ECONOMICS (cash basis: AI revenue - AI capex - AI opex)", span=4)
    rev25, rev26 = ai_rev
    put(ws, 32, 1, "AI revenue ($B)")
    put(ws, 32, 2, rev25, fmt="#,##0.0", fill=INPUT_FILL, border=True)
    put(ws, 32, 3, rev26, fmt="#,##0.0", fill=INPUT_FILL, border=True)
    put(
        ws,
        32,
        4,
        "ESTIMATE (see Methodology). MSFT $37B & Amazon $15B run-rates disclosed; Google/Meta/SpaceX estimated.",
        wrap=True,
    )
    aer = [
        (
            33,
            "AI capex ($B)",
            "=B10",
            "=C10",
            "#,##0.0",
            "= AI-infra capex (full: accel + buildings + power + net)",
            False,
        ),
        (
            34,
            "AI opex ($B)",
            "=B20/1000",
            "=C20/1000",
            "#,##0.0",
            "= annual power/operating",
            False,
        ),
        (
            35,
            "Net AI NOW ($B)",
            "=B32-B33-B34",
            "=C32-C33-C34",
            "#,##0.0",
            "revenue - capex - opex (cash burn)",
            True,
        ),
        (
            36,
            "Spend cut, our arch ($B)",
            "=B27",
            "=C27",
            "#,##0.0",
            "= accel capex avoided + opex saved",
            False,
        ),
        (
            37,
            "Net AI WITH our arch ($B)",
            "=B35+B36",
            "=C35+C36",
            "#,##0.0",
            "= Net AI now + spend cut",
            True,
        ),
        (
            38,
            "% AI spend reduction",
            "=B36/(B33+B34)",
            "=C36/(C33+C34)",
            "0%",
            "spend cut / total AI spend (accel-only; upside if DC scales)",
            False,
        ),
    ]
    for r, lab, bf, cf, fmt, note, bold in aer:
        put(ws, r, 1, lab, bold=bold)
        put(ws, r, 2, bf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 3, cf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(ws, r, 4, note, wrap=True)

    if sources:
        header(ws, 40, "SOURCES & REFERENCES", span=4)
        for i, (label, url) in enumerate(sources):
            r = 41 + i
            put(ws, r, 1, label, wrap=True)
            link = put(ws, r, 2, url)
            link.hyperlink = url
            link.font = Font(color="0563C1", underline="single")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)


def build_totals(tot, tabs):
    widths(tot, {"A": 17, "B": 12, "C": 12, "D": 11, "E": 12, "F": 13, "G": 13, "H": 9})
    header(
        tot,
        1,
        f"{len(tabs)}-COMPANY AI ECONOMICS — current AI cash burn & the effect of our architecture",
        span=8,
    )
    put(
        tot,
        2,
        1,
        "They all spend far more on AI (capex + opex) than they earn from it -- they are losing money on AI today. Our architecture cuts the accelerator spend ~95%, shrinking the burn. The named firms are a floor; a GLOBAL estimate grosses up for the rest. Yellow assumption cells (also sliders in the app) and green disclosed-data cells live on company tabs.",
        wrap=True,
    )
    tot.merge_cells("A2:H2")

    header(
        tot,
        4,
        "NET AI ECONOMICS — FY2025 (cash basis: AI revenue - AI capex - AI opex)",
        span=8,
    )
    heads = [
        "Company",
        "AI revenue ($B)",
        "AI capex ($B)",
        "AI opex ($B)",
        "Net AI NOW ($B)",
        "Spend cut, our arch ($B)",
        "Net AI W/ ARCH ($B)",
        "% spend cut",
    ]
    for j, h in enumerate(heads):
        put(tot, 5, 1 + j, h, bold=True, wrap=True)
    r0 = 6
    for k, t in enumerate(tabs):
        r = r0 + k
        put(tot, r, 1, t)
        put(tot, r, 2, f"={t}!{K_AIREV}", fmt="#,##0.0", fill=CALC_FILL, border=True)
        put(tot, r, 3, f"={t}!{K_AICAPEX}", fmt="#,##0.0", fill=CALC_FILL, border=True)
        put(tot, r, 4, f"={t}!{K_AIOPEX}", fmt="#,##0.0", fill=CALC_FILL, border=True)
        put(
            tot,
            r,
            5,
            f"={t}!{K_AINOW}",
            fmt="#,##0.0",
            fill=CALC_FILL,
            border=True,
            bold=True,
        )
        put(tot, r, 6, f"={t}!{K_AICUT}", fmt="#,##0.0", fill=CALC_FILL, border=True)
        put(
            tot,
            r,
            7,
            f"={t}!{K_AIARCH}",
            fmt="#,##0.0",
            fill=CALC_FILL,
            border=True,
            bold=True,
        )
        put(tot, r, 8, f"={t}!{K_AIPCT}", fmt="0%", fill=CALC_FILL, border=True)
    rt = r0 + len(tabs)
    put(tot, rt, 1, f"TOTAL ({len(tabs)})", bold=True)
    for col, L in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")):
        put(
            tot,
            rt,
            col,
            f"=SUM({L}{r0}:{L}{rt - 1})",
            fmt="#,##0",
            bold=True,
            fill=SUB_FILL,
            border=True,
        )
    put(
        tot,
        rt,
        8,
        f"=F{rt}/(C{rt}+D{rt})",
        fmt="0%",
        bold=True,
        fill=SUB_FILL,
        border=True,
    )
    fr = rt + 1
    nn26 = "SUM(" + ",".join(f"{t}!$C$35" for t in tabs) + ")"
    cut26 = "SUM(" + ",".join(f"{t}!$C$36" for t in tabs) + ")"
    arch26 = "SUM(" + ",".join(f"{t}!$C$37" for t in tabs) + ")"
    put(tot, fr, 1, "FY2026 (estimate)", bold=True)
    put(tot, fr, 5, f"={nn26}", fmt="#,##0", border=True, fill=SUB_FILL)
    put(tot, fr, 6, f"={cut26}", fmt="#,##0", border=True, fill=SUB_FILL)
    put(tot, fr, 7, f"={arch26}", fmt="#,##0", border=True, fill=SUB_FILL)

    sb = fr + 2
    header(tot, sb, "SAVINGS BREAKDOWN — the spend cut, split", span=8)
    put(tot, sb + 1, 1, "Cost-weighted reduction (Amdahl)")
    put(tot, sb + 1, 2, f"={RED}", fmt="0.0", fill=CALC_FILL, border=True, bold=True)
    put(tot, sb + 1, 3, "x")
    put(tot, sb + 3, 2, "Saved OPEX", bold=True)
    put(tot, sb + 3, 3, "Avoided CAPEX (Overspend)", bold=True, wrap=True)
    put(tot, sb + 3, 4, "Total", bold=True)
    opex25 = "SUM(" + ",".join(f"{t}!B26" for t in tabs) + ")/1000"
    capex25 = "SUM(" + ",".join(f"{t}!B25" for t in tabs) + ")"
    opex26 = "SUM(" + ",".join(f"{t}!C26" for t in tabs) + ")/1000"
    capex26 = "SUM(" + ",".join(f"{t}!C25" for t in tabs) + ")"
    vrows = [
        ("FY2025 annual ($B/yr)", f"={opex25}", f"={capex25}", "#,##0.0"),
        (
            "FY2025 capitalized ($B)",
            f"=B{sb + 4}/{DISC}",
            f"=C{sb + 4}/{DISC}",
            "#,##0",
        ),
        ("FY2026 annual ($B/yr)", f"={opex26}", f"={capex26}", "#,##0.0"),
        (
            "FY2026 capitalized ($B)",
            f"=B{sb + 6}/{DISC}",
            f"=C{sb + 6}/{DISC}",
            "#,##0",
        ),
    ]
    for k, (lab, bf, cf, fmt) in enumerate(vrows):
        r = sb + 4 + k
        bold = "capitalized" in lab
        put(tot, r, 1, lab, bold=bold)
        put(tot, r, 2, bf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(tot, r, 3, cf, fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
        put(tot, r, 4, f"=B{r}+C{r}", fmt=fmt, fill=CALC_FILL, border=True, bold=bold)
    gr = sb + 8
    put(tot, gr, 1, "GLOBAL capitalized -- est. ($B)", bold=True)
    put(
        tot,
        gr,
        2,
        f"=D{sb + 5}/Inputs!$B$14",
        fmt="#,##0",
        fill=CALC_FILL,
        border=True,
        bold=True,
    )
    put(
        tot,
        gr,
        3,
        f"=D{sb + 7}/Inputs!$B$14",
        fmt="#,##0",
        fill=CALC_FILL,
        border=True,
        bold=True,
    )
    put(
        tot,
        gr,
        4,
        "named-floor capitalized / named-share-of-global (Inputs B14). ESTIMATE: grosses up for other clouds, China, neoclouds, xAI, sovereign.",
        wrap=True,
    )
    pr = sb + 9
    put(tot, pr, 1, "% reduction")
    put(tot, pr, 2, f"=1-1/{OPXRED}", fmt="0%", fill=CALC_FILL, border=True)
    put(tot, pr, 3, f"=1-1/{RED}", fmt="0%", fill=CALC_FILL, border=True)
    note_r = pr + 1
    put(
        tot,
        note_r,
        1,
        "OPEX = power saved each year (recoupable). CAPEX 'Overspend' = AI capex made unnecessary. TOGGLE: Inputs 'Datacenter scaling factor' (0 = accelerator-only/conservative; 1 = whole datacenter scales with the smaller fleet). At 0 the burn ~halves; at ~0.7 net AI hits breakeven; at 1 it flips positive. Capitalized = annual / discount rate.",
        wrap=True,
    )
    tot.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=8)

    gb = note_r + 2
    header(tot, gb, "TABS", span=8)
    guide = [
        (
            " / ".join(tabs),
            "one tab each: full build capex -> infra -> servers -> accelerator -> fleet/opex -> value",
        ),
        ("Inputs", "global assumptions + the reduction engine (cost-weighted factor)"),
        (
            "Sensitivity",
            "SpaceX value across reduction tiers (10x/30x/100x/cost-weighted)",
        ),
        ("CostLadder", "own-silicon vs buy-NVIDIA vs rent $/GPU-hr"),
        (
            "Evidence",
            "BOM cost split, accelerator-share data, own-silicon TCO, filing top-lines",
        ),
        ("Methodology", "step-by-step logic, caveats, sources"),
    ]
    for k, (n, d) in enumerate(guide):
        r = gb + 1 + k
        put(tot, r, 1, n, bold=True)
        put(tot, r, 2, d, wrap=True)
        tot.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    lr = gb + 1 + len(guide) + 1
    put(
        tot,
        lr,
        1,
        "Legend: yellow = assumption / lever (each maps to a slider in the Streamlit app)  ·  green = disclosed filing or market data  ·  blue = derived formula.",
        bold=True,
    )
    tot.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=8)


def build_sensitivity(sens):
    widths(sens, {"A": 30, "B": 16, "C": 16, "D": 16, "E": 18})
    header(
        sens,
        1,
        "SENSITIVITY (SpaceX) — TODAY vs CEILING vs live cost-weighted",
        span=5,
    )
    # base = ACCELERATOR capex (matches the SpaceX tab, which reduces only accelerator silicon
    # at the conservative dc_scale=0; NOT total capex, which would overstate the avoided spend)
    SXCAP, SXOPX = "SpaceX!$B$12", "SpaceX!$B$26"
    cols = [
        ("TODAY (~20x)", "1/(Inputs!$B$4/Inputs!$B$2+(1-Inputs!$B$4)/Inputs!$B$24)"),
        ("CEILING (~41x)", "1/(Inputs!$B$4/Inputs!$B$2+(1-Inputs!$B$4)/Inputs!$B$25)"),
        ("Cost-weighted (live)", RED),
    ]
    put(sens, 2, 1, "Metric", bold=True)
    for j, (name, _) in enumerate(cols):
        put(sens, 2, 2 + j, name, bold=True, wrap=True)

    def srow(rr, lab, fmt, make):
        put(sens, rr, 1, lab)
        for j, (_, e) in enumerate(cols):
            put(sens, rr, 2 + j, make(e), fmt=fmt, fill=CALC_FILL, border=True)

    srow(3, "Efficient acquisition ($B)", "0.00", lambda e: f"={SXCAP}/({e})")
    srow(4, "Capex avoided/yr ($B)", "0.00", lambda e: f"={SXCAP}-{SXCAP}/({e})")
    srow(5, "Annual opex savings ($M)", "#,##0.0", lambda e: f"={SXOPX}")
    srow(
        6,
        "Sustained annual benefit ($B)",
        "0.00",
        lambda e: f"=({SXCAP}-{SXCAP}/({e}))+{SXOPX}/1000",
    )
    srow(
        7,
        "Capitalized value ($B)",
        "#,##0",
        lambda e: f"=(({SXCAP}-{SXCAP}/({e}))+{SXOPX}/1000)/{DISC}",
    )
    srow(
        8,
        "% of market cap",
        "0.0%",
        lambda e: f"=((({SXCAP}-{SXCAP}/({e}))+{SXOPX}/1000)/{DISC})/{MCAP}",
    )
    put(
        sens,
        10,
        1,
        "Base = SpaceX accelerator capex (SpaceX!B12), not total capex. The 'Cost-weighted' "
        "column ties to the SpaceX tab (B25/B27) at the conservative dc_scale=0.",
        wrap=True,
    )
    sens.merge_cells("A10:E10")


def build_ladder(ladder):
    widths(ladder, {"A": 34, "B": 11, "C": 11, "D": 62})
    header(
        ladder,
        1,
        "COST LADDER — $/H100-equivalent GPU-hour (at scale). Green = market price; yellow = assumption.",
        span=4,
    )
    put(ladder, 2, 1, "Procurement mode", bold=True)
    put(ladder, 2, 2, "$/hr low", bold=True)
    put(ladder, 2, 3, "$/hr high", bold=True)
    put(ladder, 2, 4, "What's baked into the price", bold=True)
    lad = [
        (
            "Own custom silicon (TPU/Trainium)",
            0.9,
            1.4,
            "COGS + modest Broadcom/Marvell margin + power + DC. NO NVIDIA margin.",
        ),
        (
            "Buy + operate NVIDIA (scale)",
            1.5,
            2.0,
            "NVIDIA ~84% gross margin baked into capex + power + DC.",
        ),
        (
            "Rent NVIDIA - neocloud / committed",
            2.0,
            3.5,
            "+ cloud provider capex recovery & margin.",
        ),
        (
            "Rent NVIDIA - hyperscaler on-demand",
            3.0,
            7.0,
            "+ utilization risk + flexibility premium (new B200 to ~$14).",
        ),
    ]
    for k, (m, lo, hi, note) in enumerate(lad):
        rr = 3 + k
        put(ladder, rr, 1, m)
        put(
            ladder, rr, 2, lo, fmt="0.00", fill=DATA_FILL, border=True
        )  # market-observed price
        put(ladder, rr, 3, hi, fmt="0.00", fill=DATA_FILL, border=True)
        put(ladder, rr, 4, note, wrap=True)
    header(ladder, 8, "OWNED-NVIDIA TCO CROSS-CHECK (from Inputs)", span=4)
    put(ladder, 9, 1, "Utilization")
    put(ladder, 9, 2, 0.85, fmt="0%", fill=INPUT_FILL, border=True)
    put(ladder, 10, 1, "Capex $/hr")
    put(
        ladder,
        10,
        2,
        f"={GPUCOST}/({LIFE}*8760*B9)",
        fmt="0.00",
        fill=CALC_FILL,
        border=True,
    )
    put(ladder, 10, 4, "fully-loaded $/GPU / (life x 8760h x utilization)", wrap=True)
    put(ladder, 11, 1, "Power $/hr")
    put(ladder, 11, 2, f"={PWR}*{ELEC}", fmt="0.00", fill=CALC_FILL, border=True)
    put(ladder, 11, 4, "wall kW x $/kWh", wrap=True)
    put(ladder, 12, 1, "DC/staff adder $/hr")
    put(ladder, 12, 2, 0.30, fmt="0.00", fill=INPUT_FILL, border=True)
    put(ladder, 13, 1, "Owned TCO $/hr", bold=True)
    put(
        ladder,
        13,
        2,
        "=B10+B11+B12",
        fmt="0.00",
        fill=CALC_FILL,
        border=True,
        bold=True,
    )
    put(ladder, 13, 4, "cross-checks the 'Buy + operate NVIDIA' row", wrap=True)
    nrow = 15
    for t in [
        "MARGIN STACK: own-silicon -> buy-NVIDIA ~1.4-2x (NVIDIA margin). buy -> rent ~2-3.5x (cloud margin). own -> rent ~3-5x.",
        "B200 builds ~$6,400, sells ~$40,000 -> ~84% gross margin. Hyperscalers charge 3-6x neocloud rates for identical HW.",
        "OWN-SILICON: TPU/Trainium/MTIA cost ~1/3 less per useful FLOP than NVIDIA. The model values compute at each company's ACTUAL cost (no gross-up).",
        "SemiAnalysis: TPU 20-50% lower TCO per useful FLOP vs GB200/GB300; Trainium3 ~30% better vs GB300.",
        "Sources: Spheron/IntuitionLabs (rental), Silicon Analysts (B200 cost/margin), SemiAnalysis (TPU/Trainium TCO).",
    ]:
        put(ladder, nrow, 1, t, wrap=True)
        ladder.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=4)
        nrow += 1


def build_evidence(ev):
    widths(ev, {"A": 40, "B": 14, "C": 12, "D": 52})
    header(
        ev,
        1,
        "EVIDENCE — cost split, accelerator share, own-silicon TCO, filing data",
        span=4,
    )

    def erow(rr, a, b="", c="", d="", bold=False):
        put(ev, rr, 1, a, bold=bold)
        put(ev, rr, 2, b)
        put(ev, rr, 3, c)
        put(ev, rr, 4, d, wrap=True)

    erow(
        3,
        "GPU cost split (BOM)",
        "$ cost",
        "% COGS",
        "Method 1: teardown = true resource cost",
        bold=True,
    )
    erow(4, "H100: HBM3 memory (80GB)", 1350, "41%", "MEMORY")
    erow(5, "H100: CoWoS packaging", 750, "23%", "mostly memory (interposer hosts HBM)")
    erow(6, "H100: test & assembly", 920, "28%", "shared")
    erow(7, "H100: logic die (compute)", 300, "9%", "COMPUTE -- cheapest part")
    erow(8, "H100 total COGS", 3320, "100%", "sells ~$28k -> ~88% margin")
    erow(
        9,
        "B200 total COGS",
        6400,
        "HBM 45%",
        "memory > logic die; sells ~$40k -> ~84% margin",
    )
    erow(
        11,
        "Accelerator share of server BOM",
        "",
        "",
        "for accel-within-server share",
        bold=True,
    )
    erow(12, "8x H100 server (J.P. Morgan)", "83%", "", "accelerator = $200k of $240k")
    erow(
        13,
        "8x A100 server (J.P. Morgan)",
        "71%",
        "",
        "GB200 rack ~76-80% (SemiAnalysis)",
    )
    erow(
        15,
        "Own-silicon vs NVIDIA (TCO)",
        "",
        "",
        "cheaper, but valued at cost in model",
        bold=True,
    )
    erow(
        16,
        "Google TPU v7 vs GB200/GB300",
        "20-50% lower",
        "",
        "per useful FLOP (SemiAnalysis)",
    )
    erow(
        17,
        "Amazon Trainium3 vs GB300",
        "~30% better",
        "",
        "chips ~1/3 cheaper to build",
    )
    erow(
        19,
        "Filing data (FY2025 actuals)",
        "total capex",
        "server life",
        "DISCLOSED top-lines",
        bold=True,
    )
    erow(
        20,
        "Microsoft (Jun'25, incl leases)",
        "~$88B",
        "2-6 yr",
        "'roughly half' short-lived (CFO)",
    )
    erow(21, "Alphabet", "$91.4B", "6 yr", "60% servers / 40% DC (CFO)")
    erow(
        22,
        "Amazon (cash capex)",
        "$128.3B",
        "5 yr (cut)",
        "AWS 67.8% of net P&E additions",
    )
    erow(
        23,
        "Meta (incl finance leases)",
        "$72.2B",
        "5.5 yr",
        "servers 'largest portion' (CFO)",
    )
    erow(
        24,
        "FY26 capex guidance",
        "",
        "",
        "MSFT ~$190B, Alphabet $180-190B, Amazon ~$200B, Meta $125-145B, Oracle ~$50B; SpaceX ~$18B (est)",
    )
    erow(
        26,
        "Cost-weighted reduction vs memory share",
        "reduction (x)",
        "",
        "live: =1/(w/memfac+(1-w)/flopfac)",
        bold=True,
    )
    for k, w in enumerate([0.45, 0.50, 0.60, 0.70, 0.82]):
        rr = 27 + k
        put(ev, rr, 1, f"memory share = {int(w * 100)}%")
        put(ev, rr, 2, w, fmt="0%", fill=INPUT_FILL, border=True)
        put(
            ev,
            rr,
            3,
            f"=1/(B{rr}/{MEMFAC}+(1-B{rr})/{FLOPFAC})",
            fmt="0.0",
            fill=CALC_FILL,
            border=True,
        )
        put(ev, rr, 4, "compute term dominates -> stays far below 100x")


def build_methodology(meth):
    widths(meth, {"A": 120})
    lines = [
        ("METHODOLOGY & SOURCES", True),
        ("", False),
        (
            "Engine: GPU cost ~60% memory / ~40% compute. TODAY (default): memory x100 + FLOPs x9.24 (blend x2.19 x equal-quality parameter ratio x4.22 at 1T) -> ~20x cost-weighted (Inputs B20). CEILING: FLOPs x22.1 -> ~41x.",
            False,
        ),
        (
            "  Floored by the least-reduced part (compute). Neither alone helps (mem-only ~2.5x, FLOP-only ~1.5x).",
            False,
        ),
        ("", False),
        (
            "Per company (own tab): total capex (DISCLOSED) x infra share x server share x accelerator share = accelerator capex; then fleet -> opex -> efficient -> value. FY25 + FY26.",
            False,
        ),
        (
            "  infra share strips non-AI (Amazon AWS 68%); server share CFO-disclosed (MSFT ~50%, Google 60%); accel-within-server ~67-80% from BOM.",
            False,
        ),
        (
            "  Accel % of total varies: Amazon ~30% (legacy fleet + non-AI), MSFT ~37%, Google ~40%, Meta ~48%, SpaceX ~79% (greenfield).",
            False,
        ),
        (
            "Totals (front page): rolls up the company tabs live; no double-count (they don't pay each other for the bulk). GLOBAL row grosses the named total up to a worldwide estimate (Inputs 'Named share of global AI capex').",
            False,
        ),
        (
            "Value: spend cut/yr = accel x (1 - 1/reduction) ~95% + opex savings; capitalized = annual / discount rate. Compute valued at ACTUAL cost.",
            False,
        ),
        (
            "Energy/opex reduction: DERIVED = cost-weighted reduction (energy splits memory/compute like cost). Inputs B5 defaults to =B20; overtype that cell to override.",
            False,
        ),
        (
            "HOW EACH INPUT IS DERIVED: Inputs sheet column D notes every global assumption; Evidence sheet has the BOM cost split (mem share), accelerator-share teardowns, own-silicon TCO and filing top-lines; per-company capex/shares/revenue carry a basis + clickable Sources on each company tab.",
            False,
        ),
        (
            "Net AI economics (Totals + company tabs): Net AI = AI revenue - AI capex (full AI-infra) - AI opex (cash basis). With our arch: Net AI + spend cut.",
            False,
        ),
        (
            "  Shows they all LOSE money on AI today. TOGGLE 'Datacenter scaling factor' (Inputs B13): 0 = accelerator-only (conservative); 1 = whole datacenter scales -> net AI flips positive (~0.7 = breakeven).",
            False,
        ),
        ("", False),
        ("KEY RESULTS (defaults)", True),
        (
            "- Cost-weighted reduction ~20x TODAY, ~41x at the CEILING (NOT 100x; ~16-36x over memory share 45-82% at the Today lever).",
            False,
        ),
        (
            "- Net AI FY25: 6 named firms spend ~$375B (capex+opex) vs ~$79B AI revenue = ~ -$295B/yr cash burn.",
            False,
        ),
        (
            "- With our architecture: spend cut ~$159B -> burn shrinks to ~ -$136B/yr (~42% of AI spend cut).",
            False,
        ),
        (
            "- Spend-cut value (named floor, 6% discount rate): FY25 ~$159B/yr (~$2.6T capitalized); FY26 r/r ~$337B/yr (~$5.6T). Ceiling: ~$163B FY25.",
            False,
        ),
        (
            "- GLOBAL estimate (named ~80% of world AI capex): FY25 ~$3.3T, FY26 ~$7.0T capitalized. Clearly an estimate.",
            False,
        ),
        ("", False),
        ("CAVEATS", True),
        (
            "- AI REVENUE is the softest input: MSFT $37B & Amazon $15B run-rates are DISCLOSED; Google/Meta/Oracle/SpaceX are ESTIMATES. Meta's real AI payoff is indirect ad-uplift (~$20B), not direct revenue -- so its 'loss' here overstates.",
            False,
        ),
        (
            "- Net AI is CASH basis (capex not depreciated). On an accounting (depreciation) basis the loss is smaller; on a depreciation basis it is closer to a true P&L.",
            False,
        ),
        (
            "- No company reports 'accelerator capex'. Totals DISCLOSED; server/accel split ESTIMATED (server CFO-disclosed; accel-within from BOM). +/-15-20%.",
            False,
        ),
        (
            "- FY26 is a full live chain (FY26 total-capex guidance x FY26 shares). SpaceX FY26 ~$18B is an estimate.",
            False,
        ),
        (
            "- Hyperscaler market caps (company tab row 7) are approximate placeholders -- edit to current.",
            False,
        ),
        (
            "- Own-silicon (TPU/Trainium/MTIA) is cheaper per FLOP, but compute is valued at ACTUAL cost; CostLadder shows the buy/rent premium.",
            False,
        ),
        (
            "- Multiplying the levers (100x * 9.2x ~ 900x) is NOT physical: cost is additive (Amdahl), not multiplicative.",
            False,
        ),
        (
            "- JEVONS: savings reinvested into more AI, not budget cuts. Capitalization is a simple perpetuity.",
            False,
        ),
        ("", False),
        ("SOURCES", True),
        (
            "SpaceX S-1 / IPO: https://www.hl.co.uk/news/inside-spacexs-ipo-filing-revenue-starlink-ai-and-key-financials",
            False,
        ),
        (
            "Microsoft FY25 capex (Q3 FY26 call): https://www.fool.com/earnings/call-transcripts/2026/04/29/microsoft-msft-q3-2026-earnings-transcript/",
            False,
        ),
        (
            "Alphabet FY25 $91.4B, 60/40 split (Q4'25 call): https://www.fool.com/earnings/call-transcripts/2026/02/04/alphabet-googl-q4-2025-earnings-call-transcript/",
            False,
        ),
        (
            "Amazon FY25 $128.3B + server life 6->5yr (10-K): https://www.sec.gov/Archives/edgar/data/1018724/000101872426000004/amzn-20251231.htm",
            False,
        ),
        (
            "Meta FY25 $72.2B (Q4/FY25 release): https://investor.atmeta.com/investor-news/press-release-details/2026/Meta-Reports-Fourth-Quarter-and-Full-Year-2025-Results/default.aspx",
            False,
        ),
        (
            "H100/B200 BOM + margin (Silicon Analysts): https://siliconanalysts.com/analysis/nvidia-b200-blackwell-cost-breakdown",
            False,
        ),
        (
            "TPU/Trainium TCO (SemiAnalysis): https://newsletter.semianalysis.com/p/tpuv7-google-takes-a-swing-at-the",
            False,
        ),
        (
            "GPU rental prices 2026 (Spheron): https://www.spheron.network/blog/gpu-cloud-pricing-comparison-2026/",
            False,
        ),
        (
            "Microsoft $37B AI run-rate (Q3 FY26): https://news.alphastreet.com/microsoft-msft-q3-fy2026-azure-hits-40-growth-as-ai-business-reaches-37-billion-run-rate/",
            False,
        ),
        (
            "Amazon >$15B AWS AI run-rate (Q1 FY26): https://www.bnnbloomberg.ca/business/artificial-intelligence/2026/04/09/amazon-cloud-units-ai-revenue-run-rate-exceeds-us15-billion-in-first-quarter-ceo-says/",
            False,
        ),
        (
            "Hyperscalers losing money on AI (capex vs revenue): https://fortune.com/2026/04/15/data-centers-hyperscalers-spending-billions-on-hardware-thats-worthless-in-3-years/",
            False,
        ),
    ]
    for i, (t, b) in enumerate(lines, start=1):
        put(meth, i, 1, t, bold=b, wrap=True)


def build_serving_training(ws):
    """The 2026-08-07/08 multi-GPU receipts, priced on the CostLadder rates.
    Values are computed live from ai_capex_model (MEASURED / serving_economics /
    training_throughput_ratio); every measured cell names its derived.json key
    in ai_capex_model.MEASURED. Green = measured, yellow = assumption/projection,
    blue = derived."""
    from ai_capex_model import (MEASURED, SERVING, SERVING_TRAINING_ASSUMPTIONS,
                                PREFILL_BF16_D2048_SWEEP, PREFILL_FP32_D2048_SWEEP,
                                PREFILL_TOKENS_PER_S_D2048, KV_MB_PER_TOKEN_PER_STREAM,
                                TRAIN_ADVANTAGE_SHORT_SEQ,
                                serving_cost_curve, serving_economics,
                                training_throughput_ratio)

    widths(ws, {"A": 44, "B": 42, "C": 30, "D": 14, "E": 12, "F": 60})
    header(ws, 1,
           "SERVING & TRAINING — measured 2026-08-07/08 (2/4/8 x H100), priced at the CostLadder rent mid "
           f"(${SERVING['gpu_hr']:.2f}/GPU-hr). Source: bdm/docs/deck/build/derived.json (keys in ai_capex_model.MEASURED).")
    # --- LIVE workload block ------------------------------------------------
    # Fixed rows on purpose: Inputs!B3 (flop_factor) is a formula pointing at
    # B28 below, so editing B7/B8/B9 recomputes the ENTIRE workbook the same way
    # the sidebar does in the Streamlit app.
    header(ws, 4, "WORKLOAD — edit the yellow cells; every sheet recomputes")
    put(ws, 5, 1,
        "Training, prefill and decode point in opposite directions at today's kernel maturity, so the "
        "compute lever is blended over the workload rather than fixed. Cost per GENERATED token = "
        "in:out input tokens through prefill + one token through decode. Transformer decode follows "
        "the MEASURED 1/context law at its own KV ceiling (granted no KV compression by default; "
        "raise the KV-compression cell below to grant it a mature stack).",
        wrap=True)
    put(ws, 7, 1, "Input : output token ratio")
    put(ws, 7, 2, 10.0, "0.0", fill=INPUT_FILL, border=True)
    put(ws, 7, 6, "~10:1 = code / reasoning / agent traces. 50-100:1 = RAG and document QA.", wrap=True)
    put(ws, 8, 1, "E[context] (tokens)")
    put(ws, 8, 2, 65536, "#,##0", fill=INPUT_FILL, border=True)
    put(ws, 8, 6, "Our decode cost is context-flat and the transformer's is linear, so the decode advantage is linear in this.", wrap=True)
    put(ws, 9, 1, "Training share of accelerator cost")
    put(ws, 9, 2, 0.0, "0.00", fill=INPUT_FILL, border=True)
    put(ws, 9, 6, "0 = a serving-cost claim. Training is a LOSS at short sequence today; above ~0.06 the blend drops below 1.", wrap=True)

    put(ws, 11, 1, "Measured constants", bold=True)
    put(ws, 12, 1, "Owned decode, tokens/s/box (context-flat)")
    put(ws, 12, 2, MEASURED["serve_tokens_per_s_box"], "#,##0", fill=DATA_FILL, border=True)
    put(ws, 13, 1, "KV MB per token per stream")
    put(ws, 13, 2, KV_MB_PER_TOKEN_PER_STREAM, "0.000000", fill=DATA_FILL, border=True)
    put(ws, 14, 1, "HBM TB/s per GPU")
    put(ws, 14, 2, SERVING["hbm_tbps"], "0.00", fill=INPUT_FILL, border=True)
    put(ws, 15, 1, "Transformer KV compression (mature stack)")
    put(ws, 15, 2, SERVING["tf_kv_compression"], "0.0", fill=INPUT_FILL, border=True)
    put(ws, 16, 1, "GPUs per box")
    put(ws, 16, 2, SERVING["gpus_per_box"], "0", fill=INPUT_FILL, border=True)
    put(ws, 17, 1, "Training advantage, short sequence (owned / transformer)")
    put(ws, 17, 2, TRAIN_ADVANTAGE_SHORT_SEQ, "0.000", fill=DATA_FILL, border=True)
    put(ws, 17, 6, "Transformer is 9.6x faster at T=2048 blocks (d1152). Crossover near T~88,000.", wrap=True)

    put(ws, 19, 1, "Derived", bold=True)
    put(ws, 20, 1, "KV GB per stream")
    put(ws, 20, 2, "=$B$8*$B$13/1024", "0.00", fill=CALC_FILL, border=True)
    put(ws, 21, 1, "KV GB per stream, compressed")
    put(ws, 21, 2, "=$B$20/$B$15", "0.000", fill=CALC_FILL, border=True)
    put(ws, 22, 1, "Transformer decode, tokens/s/box")
    put(ws, 22, 2, "=511.8416*32768/$B$8*$B$15*$B$16", "#,##0", fill=CALC_FILL, border=True)
    put(ws, 22, 6,
        "MEASURED 1/context law (2026-08-14 re-base): 511.84 tok/s/GPU at 32,768 ctx x 32768/E[context] "
        "x KV-compression x GPUs. Crosses our flat 562/GPU near ~30k context.", wrap=True)
    # Log-log interpolation over the measured prefill table at rows 33..39.
    put(ws, 23, 1, "Prefill table row (lookup)")
    put(ws, 23, 2, "=MATCH($B$8,$A$33:$A$39,1)", "0", fill=CALC_FILL, border=True)
    _interp = (
        "=IF(INDEX({col}$33:{col}$39,MIN($B$23+1,7))=INDEX({col}$33:{col}$39,$B$23),"
        "INDEX({col}$33:{col}$39,$B$23),"
        "EXP(LN(INDEX({col}$33:{col}$39,$B$23))"
        "+(LN($B$8)-LN(INDEX($A$33:$A$39,$B$23)))"
        "/(LN(INDEX($A$33:$A$39,MIN($B$23+1,7)))-LN(INDEX($A$33:$A$39,$B$23)))"
        "*(LN(INDEX({col}$33:{col}$39,MIN($B$23+1,7)))-LN(INDEX({col}$33:{col}$39,$B$23)))))*$B$16"
    )
    put(ws, 24, 1, "Prefill tokens/s/box — transformer")
    put(ws, 24, 2, _interp.format(col="$B"), "#,##0", fill=CALC_FILL, border=True)
    put(ws, 25, 1, "Prefill tokens/s/box — owned")
    put(ws, 25, 2, _interp.format(col="$C"), "#,##0", fill=CALC_FILL, border=True)
    put(ws, 26, 1, "Transformer seconds per generated token")
    put(ws, 26, 2, "=$B$7/$B$24+1/$B$22", "0.000000", fill=CALC_FILL, border=True)
    put(ws, 27, 1, "Owned seconds per generated token")
    put(ws, 27, 2, "=$B$7/$B$25+1/$B$12", "0.000000", fill=CALC_FILL, border=True)
    put(ws, 28, 1, "BLENDED COMPUTE LEVER (flop_factor)", bold=True)
    put(ws, 28, 2,
        "=IF($B$9>0,1/($B$9/$B$17+(1-$B$9)/($B$26/$B$27)),$B$26/$B$27)",
        "0.00", fill=CALC_FILL, border=True, bold=True)
    put(ws, 28, 6, "Inputs!B24 (TODAY lever) = this x the equal-quality parameter ratio; Inputs!B3 follows it, so this drives the whole workbook.", wrap=True)
    put(ws, 29, 1, "  of which: prefill advantage")
    put(ws, 29, 2, "=($B$7/$B$24)/($B$7/$B$25)", "0.00", fill=CALC_FILL, border=True)
    put(ws, 30, 1, "  of which: decode advantage")
    put(ws, 30, 2, "=(1/$B$22)/(1/$B$12)", "0.0", fill=CALC_FILL, border=True)

    put(ws, 32, 1, "Measured prefill throughput, tokens/s/GPU (bf16 saturated lane, d2048)", bold=True)
    for c, t in enumerate(["Context", "Transformer", "Owned"], start=1):
        put(ws, 32, c + 3, t, bold=True, fill=SUB_FILL, border=True)
    for i, ctx in enumerate(sorted(PREFILL_TOKENS_PER_S_D2048)):
        tf_tps, own_tps = PREFILL_TOKENS_PER_S_D2048[ctx]
        put(ws, 33 + i, 1, ctx, "#,##0", border=True)
        put(ws, 33 + i, 2, tf_tps, "#,##0", fill=DATA_FILL, border=True)
        put(ws, 33 + i, 3, own_tps, "#,##0", fill=DATA_FILL, border=True)

    put(ws, 41, 1,
        "SCOPE: the serving and training blocks below are component-scope systems benches (bAttention "
        "d1536 fwd+bwd h-recurrence sub-block vs transformer d2048 forward-only layer); every speedup "
        "there is each family vs its OWN 1-GPU baseline. The prefill lanes quoted further down are "
        "cross-family: bf16 parameter-matched at block scope, and fp32 at full-model scope.",
        wrap=True)

    header(ws, 43, "Serving at long context — $/1M generated tokens, one 8xH100 box")
    for c, t in enumerate(["Context (tokens)", "bAttention $/1M tok", "Transformer $/1M tok (mature stack)",
                           "TF streams/GPU", "Cost ratio", "Note"], start=1):
        put(ws, 44, c, t, bold=True, fill=SUB_FILL, border=True)
    r = 45
    for row in serving_cost_curve():
        note = ("bAttention rate extrapolated past 262k (flatness measured 4k-262k)"
                if row["battn_extrapolated"] else
                ("a measured decode cell (64 streams/GPU against the transformer 1)" if row["ctx"] == 262144 else ""))
        put(ws, r, 1, row["ctx"], "#,##0", border=True)
        put(ws, r, 2, row["battn_usd_per_mtok"], "$0.0000",
            fill=INPUT_FILL if row["battn_extrapolated"] else DATA_FILL, border=True)
        put(ws, r, 3, row["tf_usd_per_mtok"], "$0.00", fill=CALC_FILL, border=True)
        put(ws, r, 4, row["tf_streams_per_gpu"], "0", fill=CALC_FILL, border=True)
        put(ws, r, 5, row["cost_ratio"], "0.0×", fill=CALC_FILL, border=True)
        put(ws, r, 6, note, wrap=True, border=True)
        r += 1
    put(ws, r, 1,
        "RE-BASED 2026-08-14 on the full-model decode receipt: per-GPU aggregate throughput at each family's "
        "own concurrency ceiling, transformer granted no KV compression. Ratio ~ context/29,800, so the "
        "TRANSFORMER is ahead below ~30k context; x2.2 at 64k, x8.8 at 262k. Memory-ceiling result, not a "
        "per-token one - at one stream and 64k the transformer decodes a token faster than we do. Granting it "
        f"KV /8: x{serving_economics(262144, s={'tf_kv_compression': 8.0})['cost_ratio']:.1f} at 262k.", wrap=True)
    r += 2

    header(ws, r, "Training at scale — same cluster, more steps/s")
    r += 1
    train_rows = [
        ("8-GPU speedup, matched load (16 in flight)", "x5.15 fwd+bwd vs x3.70 fwd-only Ulysses best",
         f"x{training_throughput_ratio():.2f} cluster throughput", DATA_FILL, "MEASURED — matched_load_curve"),
        ("8-GPU speedup, deep load (256 in flight)", "x6.27 (pipeline keeps filling; TF saturated by 16)",
         f"x{training_throughput_ratio(deep=True):.2f}", DATA_FILL, "MEASURED — scaling_1to8_best"),
        ("GPU-hours for the same training work", "-28% (matched) … -41% (deep)", "", CALC_FILL, "derived"),
        ("64k-token sequence on one 80 GB GPU", "30.9 GB fits vs OOM (78.4 GB attempted)", "", DATA_FILL,
         "MEASURED — pipeline_feasibility"),
        ("Pipeline per-GPU peak (flat in load)", "9.05 GB vs 43.3 GB GPipe stage", "x4.8 (component-scope, width-unmatched)",
         DATA_FILL, "MEASURED — pipeline_memory_asymmetry"),
        ("Params for equal quality at 70B", f"x{MEASURED['parity_70B_param_multiple']:.1f} [2.2-7.7] transformer",
         "", INPUT_FILL, "PROJECTION — quality_fit.ref_70B"),
        ("Tokens for equal quality at 70B", f"x{MEASURED['parity_70B_token_multiple']:.2f} [1.33-2.10] (beta=0.28 assumption)",
         "", INPUT_FILL, "PROJECTION — quality_fit.ref_70B"),
        ("Stepped vs scanned training", f"x{MEASURED['stepped_vs_scanned']:.0f} cost to step the recurrence token-by-token",
         "", DATA_FILL, "MEASURED — stepped_vs_scanned"),
    ]
    for c, t in enumerate(["Metric", "Value", "Ratio / consequence", "", "", "Status — derived.json key"], start=1):
        put(ws, r, c, t, bold=True, fill=SUB_FILL, border=True)
    r += 1
    for name, val, ratio, fill, status in train_rows:
        put(ws, r, 1, name, border=True)
        put(ws, r, 2, val, fill=fill, border=True, wrap=True)
        put(ws, r, 3, ratio, fill=CALC_FILL if ratio else None, border=True)
        put(ws, r, 6, status, border=True, wrap=True)
        r += 1
    r += 1

    header(ws, r, "The compute lever — blended over the workload (sets flop_factor)")
    r += 1
    put(ws, r, 1,
        "Training, prefill and decode point in OPPOSITE directions at today's kernel maturity, so a "
        "single number cannot represent them. Training is a loss at short sequence (transformer "
        "5.3-9.6x faster at T=2048 blocks; FLOPs near-matched at 1.07-1.19x, the gap is achieved "
        "FLOP/s 200.9 vs 21.7 TF/s; crossover near T~88,000). Prefill crosses over near 65k context. "
        "Decode crosses in our favour near ~30k context. The lever is therefore blended over the workload: per "
        "generated token, in:out input tokens through prefill plus one token through decode, using "
        "measured per-GPU throughput at each family's own concurrency ceiling, the transformer granted "
        "no KV compression. At the default operating point (10:1, 64k context, training share 0): "
        "prefill x1.17, decode x2.20, blended x2.19, with decode ~99.7% of transformer serving cost. "
        "The decode half is a MEMORY CEILING, not a per-token result - at one stream and 64k the "
        "transformer decodes a token more cheaply than we do - and it crosses 1 at ~30k context, below "
        "which the transformer wins. Granting it KV/8 divides our decode lever by 8. A training share "
        "above ~0.06 also takes the blend below 1. All are reachable in the live model.",
        wrap=True)
    r += 2
    put(ws, r, 1,
        "The prefill component itself exists in two lanes, both shown below. It comes from the bf16 lane, where "
        "both families run their production attention kernels (there is no fp32 FlashAttention kernel). "
        "BF16 LANE: parameter counts exactly matched, 1 layer, batch 16, 1 x H100 80GB HBM3, 2026-07-21 — "
        "x4.02 at d512/1M, x3.83 at d1024/1M, x2.01 at d2048/262k. Receipts: "
        "experiments/paper_figures/output/matched_d{512,1024,2048}_h100.json",
        wrap=True)
    r += 2
    for c, t in enumerate(["Context (tokens)", "Transformer ms", "bAttention ms", "TF / bAttention", "", "Lane / note"], start=1):
        put(ws, r, c, t, bold=True, fill=SUB_FILL, border=True)
    r += 1
    for label, sweep, fill in (("bf16 (lever)", PREFILL_BF16_D2048_SWEEP, DATA_FILL),
                               ("fp32 (full model)", PREFILL_FP32_D2048_SWEEP, INPUT_FILL)):
        for T_tok, tf_ms, ba_ms in sweep:
            ratio = tf_ms / ba_ms
            note = label + (" — transformer ahead" if ratio < 1.0 else "")
            put(ws, r, 1, T_tok, "#,##0", border=True)
            put(ws, r, 2, tf_ms, "#,##0.0", fill=CALC_FILL, border=True)
            put(ws, r, 3, ba_ms, "#,##0.0", fill=fill, border=True)
            put(ws, r, 4, ratio, "0.00×", fill=CALC_FILL, border=True)
            put(ws, r, 6, note, wrap=True, border=True)
            r += 1
    put(ws, r, 1,
        "Both blocks are d2048, so the lanes are directly comparable: 2.01x in bf16 against 7.91x in fp32 "
        "at 262k. The fp32 lane is full-model scope (24 layers, batch 1) and its receipt records "
        "FLASH_ATTENTION and CUDNN_ATTENTION as 'rejected: No available kernel', with the owned arm's "
        "route attestation reading dtype_mix io=fp32,bf_col=fp16; its params differ by 16.7%. Full model "
        "in bf16 has not been run. Receipt: "
        "experiments/paper_figures/output/deck_speed_scaling_d2048_h100_20260803.json", wrap=True)
    r += 2

    header(ws, r, "16-bit IO in training")
    r += 1
    put(ws, r, 1,
        "A training step-time gate: same architecture on every row, only the IO dtype varies. "
        "d1536 / 27 layers / block 2048 / batch 32 / seed 1, 1 x GH200 480GB, GPU verified idle before "
        "each arm, steady state only. 0 skipped steps on all arms; verdict PASS_io16_SHIPS. Throughput "
        "and memory rows are d1536/L27; the val deltas come from the gate's smaller vehicle (width 512, "
        "15 layers). This is why the production training lane is 16-bit; it is not an input to the "
        "compute lever above. Receipt: experiments/paper_figures/output/receipts/io16_gate_20260809.md",
        wrap=True)
    r += 2
    for c, t in enumerate(["IO dtype", "s/step", "Speed vs fp32", "Peak allocated (GB)", "Memory vs fp32", "Val cost @601 steps"], start=1):
        put(ws, r, c, t, bold=True, fill=SUB_FILL, border=True)
    r += 1
    io_rows = [
        ("fp32", MEASURED["io16_fp32_s_per_step"], 1.0,
         MEASURED["io16_fp32_peak_mb"] / 1024, 1.0, "baseline"),
        ("fp16", MEASURED["io16_fp16_s_per_step"], MEASURED["io16_fp16_speedup"],
         MEASURED["io16_fp16_peak_mb"] / 1024, MEASURED["io16_fp16_mem_ratio"], "+0.00128"),
        ("bf16", MEASURED["io16_bf16_s_per_step"], MEASURED["io16_bf16_speedup"],
         MEASURED["io16_bf16_peak_mb"] / 1024, MEASURED["io16_bf16_mem_ratio"], "+0.00391"),
    ]
    for name, sps, spd, gb, memr, val in io_rows:
        put(ws, r, 1, name, border=True)
        put(ws, r, 2, sps, "0.000", fill=DATA_FILL, border=True)
        put(ws, r, 3, spd, "0.000×", fill=CALC_FILL, border=True)
        put(ws, r, 4, gb, "0.0", fill=DATA_FILL, border=True)
        put(ws, r, 5, memr, "0.000×", fill=CALC_FILL, border=True)
        put(ws, r, 6, val, border=True)
        r += 1
    r += 1

    header(ws, r, "Assumptions — one row each (MEASURED / PROJECTION / ASSUMPTION)")
    r += 1
    for c, t in enumerate(["Assumption", "Value", "", "", "Status", "Source"], start=1):
        put(ws, r, c, t, bold=True, fill=SUB_FILL, border=True)
    r += 1
    for a, b, st_, src in SERVING_TRAINING_ASSUMPTIONS:
        put(ws, r, 1, a, border=True)
        put(ws, r, 2, b, wrap=True, border=True)
        put(ws, r, 5, st_, fill=DATA_FILL if st_ == "MEASURED" else INPUT_FILL, border=True)
        put(ws, r, 6, src, wrap=True, border=True)
        r += 1


def main() -> None:
    wb = Workbook()
    tot = wb.active
    tot.title = "Totals"
    # companies single-sourced from ai_capex_model
    tabs = [c["name"] for c in COMPANIES]
    sheets = {name: wb.create_sheet(name) for name in tabs}
    inp = wb.create_sheet("Inputs")
    sens = wb.create_sheet("Sensitivity")
    ladder = wb.create_sheet("CostLadder")
    servtrain = wb.create_sheet("ServingTraining")
    ev = wb.create_sheet("Evidence")
    meth = wb.create_sheet("Methodology")

    build_inputs(inp)
    for c in COMPANIES:
        build_company(
            sheets[c["name"]],
            c["name"],
            c["fy25"],
            c["fy26"],
            c["mcap"],
            c["ai_rev"],
            c["basis"],
            c.get("sources", ()),
        )
    build_totals(tot, tabs)
    build_sensitivity(sens)
    build_ladder(ladder)
    build_serving_training(servtrain)
    build_evidence(ev)
    build_methodology(meth)

    wb.calculation.fullCalcOnLoad = True
    out = "AI_Capex_Efficiency.xlsx"
    wb.save(out)
    print(f"wrote {out} with tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
